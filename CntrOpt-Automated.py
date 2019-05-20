#log: Added if CBM = 0, quit. to prevent keep going.


import gc
import openpyxl
from Optimizationv2 import *
import warnings

from openpyxl import Workbook

warnings.filterwarnings("ignore")


x20 = 27
x40 = 57
x40HC = 67
x45 = 0

print("Note: This program will give a 'Simulation Complete!' status if it ran successfully.")
print("      If it exit, it means the program did not complete successfully.")
print("      If multiple failure, please contact Sunny.")
print("")

print("===The following containers are being used:===")
print("Note: 0CBM = container not being used.")
print("   20' Container: ", x20, "CBM")
print("   40' Container: ", x40HC, "CBM")
print("   40HC Container: ", x40, "CBM")
print("   45' Container: ", x45, "CBM")
print("==============================================")
print("")

#x = input("Now starting looking at read.xlsx").lower()

#set up:
wb = openpyxl.load_workbook('readLite.xlsx')
sheet = wb.get_sheet_by_name('Sheet')

#x = input("Finished looking at xlsx").lower()



#Set up headers:
sheet.cell(row=1, column=3).value = "20'"
sheet.cell(row=1, column=4).value = "40'"
sheet.cell(row=1, column=5).value = "40HC"
sheet.cell(row=1, column=6).value = "45'"
sheet.cell(row=1, column=7).value = "Utilization"

#x = input("About to start doing sim").lower()

#loop and input to right.
rowCount = sheet.get_highest_row()+1


for rowNum in range(2, rowCount):  # skip the first row
    #print("Currently doing : ", rowNum)
    identifier = sheet.cell(row=rowNum, column=1).value
    CBM = sheet.cell(row=rowNum, column=2).value
    if CBM == 0:
        break
    cntrOpt = Optimization(x20, x40, x40HC, x45, CBM)
    cntrRslt = cntrOpt.LP()
    sheet.cell(row=rowNum, column=1).value = identifier
    sheet.cell(row=rowNum, column=3).value = cntrRslt[0] #20
    sheet.cell(row=rowNum, column=4).value = cntrRslt[1] #40
    sheet.cell(row=rowNum, column=5).value = cntrRslt[2] #40HC
    sheet.cell(row=rowNum, column=6).value = cntrRslt[3] #45
    sheet.cell(row=rowNum, column=7).value = cntrRslt[4] #Utilization

    #x = input("Finished with a sim").lower()
    del cntrRslt
    del cntrOpt
    gc.collect()

wb.save('updatedLite.xlsx')

#x = input("Simulation Completed!")




