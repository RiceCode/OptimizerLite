#log: Added if CBM = 0, quit. to prevent keep going.



import gc
import openpyxl
from Optimizationv2 import *
from Init import *
import warnings

from openpyxl import Workbook

warnings.filterwarnings("ignore")


initStart = Init()
initPr = initStart.Prompt()
xLcl = initPr[0]
x20min = initPr[1]
x20 = initPr[2]
x40 = initPr[3]
x40HC = initPr[4]
x45 = initPr[5]

print("Note: This program will give a 'Simulation Complete!' status if it ran successfully.")
print("      If it exit, it means the program did not complete successfully.")
print("      If multiple failure, please contact Sunny.")
print("")

print("===The following containers are being used:===")
print("Note: 0CBM = container not being used.")
print("   LCL is activated with 1 CBM each")
print("   20' Container minimum is: ", x20min, "CBM")
print("   20' Container: ", x20, "CBM")
print("   40' Container: ", x40, "CBM")
print("   40HC Container: ", x40HC, "CBM")
print("   45' Container: ", x45, "CBM")
print("==============================================")
print("")

#x = input("Now starting looking at read.xlsx").lower()

#set up:
wb = openpyxl.load_workbook('readLite.xlsx')
sheet = wb.get_sheet_by_name('Sheet')
rowCount = sheet.max_row

print(rowCount)

#x = input("Finished looking at xlsx").lower()



#Set up headers:
sheet.cell(row=1, column=2).value = "LCL"
sheet.cell(row=1, column=3).value = "20'"
sheet.cell(row=1, column=4).value = "40'"
sheet.cell(row=1, column=5).value = "40HC"
sheet.cell(row=1, column=6).value = "45'"
sheet.cell(row=1, column=7).value = "Utilization"




for rowNum in range(2, rowCount+1):  # skip the first row
    #print("Currently doing : ", rowNum)
    CBM = sheet.cell(row=rowNum, column=1).value
    if CBM == 0:
        break
    cntrOpt = Optimization(xLcl, x20min, x20, x40, x40HC, x45, CBM)
    cntrRslt = cntrOpt.LP()
    sheet.cell(row=rowNum, column=2).value = cntrRslt[0] #LCL
    sheet.cell(row=rowNum, column=3).value = cntrRslt[1] #20
    sheet.cell(row=rowNum, column=4).value = cntrRslt[2] #40
    sheet.cell(row=rowNum, column=5).value = cntrRslt[3] #40HC
    sheet.cell(row=rowNum, column=6).value = cntrRslt[4] #45
    sheet.cell(row=rowNum, column=7).value = cntrRslt[5] #Utilization

    #x = input("Finished with a sim").lower()
    del cntrRslt
    del cntrOpt
    gc.collect()

wb.save('updatedLite.xlsx')

x = input("Simulation Completed!")




